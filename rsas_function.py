from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QDir
from PyQt5.QtWidgets import QMainWindow, QMessageBox, QFileDialog
from functools import partial
import sys
import rsas_ui
import xlrd
import os
import pandas as pd
import numpy as np
import re
import chardet
import codecs
import getopt
import shutil
import sys
import subprocess
import xlrd
import os
import sys
import openpyxl

import importlib


class Show_Main(QMainWindow, rsas_ui.Ui_MainWindow):
    def __init__(self):
        super(Show_Main, self).__init__()
        self.setupUi(self)



        #选择文件夹
        #self.Connection_test_Button.clicked.connect(lambda: )
        self.file_dir.clicked.connect(lambda: self.loadDir())
        #self.xuanze_file_dir.clicked.connect(lambda: QMessageBox.about(self, "提示", "我被点击了，哈哈哈哈"))

        #excel合并的按钮
        #self.txt_to_excel.clicked.connect(lambda: txtToExcel().txt_to_excel_run())
        self.start.clicked.connect(lambda: rsas_jx().write_data(self.file_dir_show.toPlainText()))


    def loadDir(self):
        filenames = QFileDialog.getExistingDirectory(None, "选择存储文件夹",os.getcwd())  # 返回选择文件的名字
        print(filenames)
        self.file_dir_show.setPlainText(filenames)

    def tishi(self,zhuji,loudong):
        print("共完成了" + zhuji + "个主机IP的整理; " + "共计" + loudong + "个漏洞!!!")
        QMessageBox.about(self, "提示", "共完成了" + zhuji + "个主机IP的整理; " + "共计" + loudong + "个漏洞!!!"  + "\n\n详情请查阅" + dst_path + " 文件!")

    def error_not_found(self,zhuji):
        print(zhuji)
        QMessageBox.warning(self, "警告！","共完成了" + zhuji + "个主机IP的整理; " + "\n\n 请检查所选文件夹是否有误！！")

    def error_not_found2(self):
        QMessageBox.warning(self, "警告！","共完成了0个主机IP的整理; " + "\n\n 请检查所选文件夹是否有误！！")





dst_path = 'rsas_result.xlsx'

class rsas_jx():

    @classmethod
    def load_vuln(self,filename,src_path):
        ip = filename.replace(src_path, '').replace('.xls', '')
        ef = xlrd.open_workbook(filename)
        sheets = ef.sheet_names()
        result = []
        try:
            for i in sheets:
                if i == "漏洞信息":
                    sheet = ef.sheet_by_name(i)
                    lastport = ''
                    lastservice = ''
                    for r in range(1, sheet.nrows):
                        port = str(sheet.cell_value(r, 0)).replace('.0', '')
                        service = sheet.cell_value(r, 2)
                        if port != '':
                            lastport = port
                            lastservice = service
                        else:
                            port = lastport
                            service = lastservice
                        vulnname = sheet.cell_value(r, 3)
                        degree = str(sheet.cell_value(r, 5))
                        degree = degree.replace('[', '').replace(']', '')
                        vulndesc = sheet.cell_value(r, 17)
                        suggest = sheet.cell_value(r, 18)

                        # rec = ip + "|" + port + "|" + service + "|" + vulnname + "|" + degree + "|" + vulndesc + "|" + suggest
                        rec1 = ip + "|" + vulnname + "|" + vulndesc + "|" + degree + "|" + suggest
                        # result1.append(rec)
                        result.append(rec1)
                        # print(result1)

            print(ip + ' success!')
            # return result
            return result

        # except Exception, e:
        #   print e
        except Exception as ex:
            print(ex)

    @classmethod
    def openfile(self,src_path_input):

        self.src_path_input = src_path_input
        self.src_path = self.src_path_input + '/'

        f = open(dst_path, 'w')
        f.close()
        result = []
        ipgeshu = 0
        for root, dirs, files in os.walk(self.src_path):
            count = 0
            for i in files:
                ipgeshu = ipgeshu + 1
                count = count + 1
                fn = self.src_path + i
                if fn[-4:] == '.xls':
                    # print ('start ' + str(count) + ' files')
                    subresult = self.load_vuln(fn,self.src_path)
                    if subresult:
                        for j in subresult:
                            result.append(j)
        return result,ipgeshu

    @classmethod
    def add_number(self,src_path_input):
        # 添加序号，需要个result
        number = 0
        result2 = []
        openfile = self.openfile(src_path_input)
        for j in openfile[0]:
            number = number + 1
            rec = str(number) + "|" + j
            result2.append(rec)
        return result2,number,openfile[1]

    @classmethod
    def write_data(self,src_path_input):
        if src_path_input == '':
            Show_Main().error_not_found2()
        else:
            try:
                add_number = self.add_number(src_path_input)

                fl = openpyxl.Workbook()
                sheetfl = fl.active
                sheetfl.title = u'漏洞扫描'
                sheetfl.cell(1, 1).value = "序号"
                sheetfl.cell(1, 2).value = "主机ip"
                sheetfl.cell(1, 3).value = "漏洞名称"
                sheetfl.cell(1, 4).value = "漏洞描述"
                sheetfl.cell(1, 5).value = "风险等级"
                sheetfl.cell(1, 6).value = "整改建议"

                hangNumber = 2  # 让数据从第二行开始存放
                #print(add_number[0])
                for i in add_number[0]:
                    r = i.split("|")
                    sheetfl.cell(hangNumber, 1).value = r[0]
                    sheetfl.cell(hangNumber, 2).value = r[1]
                    sheetfl.cell(hangNumber, 3).value = r[2]
                    sheetfl.cell(hangNumber, 4).value = r[3]
                    sheetfl.cell(hangNumber, 5).value = r[4]
                    sheetfl.cell(hangNumber, 6).value = r[5]
                    hangNumber = hangNumber + 1
                    # print ("record " + str(count) + " done")
                #fl.save(dst_path)
                #print("共完成了" + str(add_number[2]) + "个主机IP的整理：" + "共计" + str(add_number[1]) + "个漏洞!!!")
                fl.save(dst_path)
                zhuji = str(add_number[2])
                loudong = str(add_number[1])
                if zhuji == '0':
                    Show_Main().error_not_found(zhuji)
                else:
                    Show_Main().tishi(zhuji, loudong)
            except:
                Show_Main().error_not_found2()



if __name__ == '__main__':
    #获取UIC窗口操作权限
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = Show_Main()
    # 调自定义的界面（即刚转换的.py对象）
    #Ui = ControlBoard()  # 这里也引用了一次rwap.rwap_gui.mainwindows.py文件的名字注意
    #Ui.setupUi(MainWindow)
    # 显示窗口并释放资源
    MainWindow.show()
    sys.exit(app.exec_())
