from PyQt5.QtWidgets import QFileDialog, QMessageBox
import Rsas_Ui
import xlrd
import os
import openpyxl



class Rsas_Function(Rsas_Ui.Rsas_uimian):

    def __init__(self):
        super(Rsas_Function, self).__init__()
        #sys.stdout = sys.stdout  # 打印输出
        self.rsas_ui()

    def rsas_Implement(self):
        # 选择文件夹
        self.rsas_file_btn.clicked.connect(lambda: self.rsas_LoadDir())
        # excel合并的按钮
        self.rsas_start_btn.clicked.connect(lambda: rsas_jx().write_data(self.rsas_lineEdit.text(),self.rsas_QComboBox_rules.currentText()))

    def rsas_LoadDir(self):
        filenames = QFileDialog.getExistingDirectory(None, "选择存储文件夹", os.getcwd())  # 返回选择文件的名字
        if filenames: #如果值不为空，才进行赋值，为空就不操作，就可以保留了原来的值了
            self.rsas_lineEdit.setText(filenames)





    def rsas_tishi(self,zhuji,loudong):
        #print("共完成了" + zhuji + "个主机IP的整理; " + "共计" + loudong + "个漏洞!!!")
        QMessageBox.about(self, "提示", "共完成了" + zhuji + "个主机IP的整理; " + "共计" + loudong + "个漏洞!!!"  + "\n\n详情请查阅" + dst_path + " 文件!")

    def rsas_error_not_found(self,zhuji):
        #print(zhuji)
        QMessageBox.warning(self, "警告！","共完成了" + zhuji + "个主机IP的整理; " + "\n\n 请检查所选文件夹是否有误！！")

    def rsas_error_not_found2(self):
        QMessageBox.warning(self, "警告！","共完成了0个主机IP的整理; " + "\n\n 请检查所选文件夹是否有误！！")


    def rsas_error_not_found3(self):
        QMessageBox.warning(self, "警告！","      漏洞个数为0！！\n \n可能是规则不匹配，请您检查当前规则是否匹配！！")



dst_path = 'rsas_result.xlsx'

class rsas_jx():

    @classmethod
    def load_vuln(self,filename,src_path,input_rules):
        ip = filename.replace(src_path, '').replace('.xls', '')
        ef = xlrd.open_workbook(filename)
        sheets = ef.sheet_names()
        result = []
        try:
            for i in sheets:
                #if i == "漏洞信息":
                if i == input_rules:
                    sheet = ef.sheet_by_name(i)
                    lastport = ''
                    lastservice = ''
                    if input_rules == "漏洞信息":
                        input_rules_flag = 1 #根据输入的规则，决定从第几行开始读取内容
                    else:
                        input_rules_flag = 2
                    for r in range(input_rules_flag, sheet.nrows):
                        port = str(sheet.cell_value(r, 0)).replace('.0', '')
                        service = sheet.cell_value(r, 2)
                        if port != '':
                            lastport = port
                            lastservice = service
                        else:
                            port = lastport
                            service = lastservice
                        vulnname = sheet.cell_value(r, 3)
                        degree = str(sheet.cell_value(r, 5))
                        degree = degree.replace('[', '').replace(']', '')
                        vulndesc = sheet.cell_value(r, 17)
                        suggest = sheet.cell_value(r, 18)

                        # rec = ip + "|" + port + "|" + service + "|" + vulnname + "|" + degree + "|" + vulndesc + "|" + suggest
                        rec1 = ip + "|" + vulnname + "|" + vulndesc + "|" + degree + "|" + suggest
                        # result1.append(rec)
                        result.append(rec1)
                        # print(result1)

            #print(ip + ' success!')
            # return result
            return result

        # except Exception, e:
        #   print e
        except Exception as ex:
            print(ex)

    @classmethod
    def openfile(self,src_path_input,input_rules):

        self.src_path_input = src_path_input
        self.src_path = self.src_path_input + '/'

        f = open(dst_path, 'w')
        f.close()
        result = []
        ipgeshu = 0
        for root, dirs, files in os.walk(self.src_path):
            count = 0
            for i in files:
                ipgeshu = ipgeshu + 1
                count = count + 1
                fn = self.src_path + i
                if fn[-4:] == '.xls':
                    # print ('start ' + str(count) + ' files')
                    subresult = self.load_vuln(fn,self.src_path,input_rules)
                    if subresult:
                        for j in subresult:
                            result.append(j)
        return result,ipgeshu

    @classmethod
    def add_number(self,src_path_input,input_rules):
        # 添加序号，需要个result
        number = 0
        result2 = []
        openfile = self.openfile(src_path_input,input_rules)
        for j in openfile[0]:
            number = number + 1
            rec = str(number) + "|" + j
            result2.append(rec)
        return result2,number,openfile[1]

    @classmethod
    def write_data(self,src_path_input,input_rules):
        if src_path_input == '':
            Rsas_Function().rsas_error_not_found2()
        else:
            try:
                if input_rules == '':
                    input_rules = '漏洞信息'
                else:
                    input_rules = input_rules

                add_number = self.add_number(src_path_input,input_rules)

                fl = openpyxl.Workbook()
                sheetfl = fl.active
                sheetfl.title = u'漏洞扫描'
                sheetfl.cell(1, 1).value = "序号"
                sheetfl.cell(1, 2).value = "主机ip"
                sheetfl.cell(1, 3).value = "漏洞名称"
                sheetfl.cell(1, 4).value = "漏洞描述"
                sheetfl.cell(1, 5).value = "风险等级"
                sheetfl.cell(1, 6).value = "整改建议"

                hangNumber = 2  # 让数据从第二行开始存放
                #print(add_number[0])
                for i in add_number[0]:
                    r = i.split("|")
                    sheetfl.cell(hangNumber, 1).value = r[0]
                    sheetfl.cell(hangNumber, 2).value = r[1]
                    sheetfl.cell(hangNumber, 3).value = r[2]
                    sheetfl.cell(hangNumber, 4).value = r[3]
                    sheetfl.cell(hangNumber, 5).value = r[4]
                    sheetfl.cell(hangNumber, 6).value = r[5]
                    hangNumber = hangNumber + 1
                    # print ("record " + str(count) + " done")
                #fl.save(dst_path)
                #print("共完成了" + str(add_number[2]) + "个主机IP的整理：" + "共计" + str(add_number[1]) + "个漏洞!!!")
                fl.save(dst_path)
                zhuji = str(add_number[2])
                loudong = str(add_number[1])
                if zhuji == '0':
                    Rsas_Function().rsas_error_not_found(zhuji)
                else:
                    Rsas_Function().rsas_tishi(zhuji, loudong)
                    if loudong == '0':
                        Rsas_Function().rsas_error_not_found3()
            except:
                Rsas_Function().rsas_error_not_found2()


